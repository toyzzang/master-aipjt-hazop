from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.hazop import ActionPlanRow, GuidewordRow, NodeRow, RiskAssessmentRow, RiskCriteria, RiskCriterion


NODE_SHEET = "#1 노드리스트"
GUIDEWORD_SHEET = "#2 가이드워드"
RISK_SHEET = "#3 위험성평가"
ACTION_SHEET = "#4 조치계획서"
CRITERIA_SHEET = "위험도기준"

NODE_HEADERS = ["노드순서", "노드명"]
GUIDEWORD_HEADERS = ["노드순서", "노드명", "변수", "가이드워드"]
RISK_HEADERS = [
    "No",
    "노드순서",
    "노드명",
    "변수",
    "가이드워드",
    "일탈",
    "원인",
    "결과",
    "현재안전조치",
    "빈도",
    "강도",
    "위험도",
    "위험도판단",
    "조치필요여부",
    "판단근거",
    "강도근거",
    "빈도근거",
    "비고",
]
ACTION_HEADERS = [
    "No",
    "위험성평가No",
    "노드명",
    "개선권고사항",
    "조치후빈도",
    "조치후강도",
    "조치후위험도",
    "근거",
    "비고",
]
CRITERIA_HEADERS = ["구분", "점수", "기준"]
DEFAULT_CRITERIA_ROWS = [
    ["빈도", 1, "10년에 1회 정도 발생할 경우 또는 없을 경우"],
    ["빈도", 2, "5년에 1회 정도 발생할 경우"],
    ["빈도", 3, "1년에 1회 정도 발생할 경우"],
    ["빈도", 4, "1개월 1회 정도 발생할 경우"],
    ["빈도", 5, "1일 1회 정도 발생할 경우"],
    ["강도", 1, "영향없음, 이상등급/관리사고"],
    ["강도", 2, "경미한 불휴업 재해, C급/준사고"],
    ["강도", 3, "경미한 휴업 재해, B급 경미재해"],
    ["강도", 4, "중대재해, A급 사망 등 중대재해"],
    ["위험도", "1~3", "무시할 수 있는 위험 - 현재 안전대책 유지"],
    ["위험도", "4~6", "미미한 위험 - 안전정보 및 주기적 표준작업안전 교육 필요"],
    ["위험도", 8, "경미한 위험 - 표지부착, 작업절차서 표기 등 관리적 대책 필요"],
    ["위험도", "9~11", "상당한 위험 - 계획된 정비/보수기간에 위험성 감소대책 필요"],
    ["위험도", "12~15", "중대한 위험 - 긴급 임시안전대책 후 계획 정비/보수기간에 안전대책 필요"],
    ["위험도", "16~20", "허용불가 위험 - 즉시 작업중단 및 즉시 개선 필요"],
]


def read_nodes_from_excel(path_or_file) -> list[NodeRow]:
    """업로드 Excel에서 `#1 노드리스트`만 읽습니다.

    웹 화면에서 파일을 고르자마자 Node별 물질 입력칸을 만들 때 사용합니다.
    이 단계는 "Node 이름 목록을 미리 보는 것"이 목적이라 `#2 가이드워드`는 검사하지 않습니다.
    """

    workbook = load_workbook(path_or_file, read_only=True, data_only=True)
    try:
        if NODE_SHEET not in workbook.sheetnames:
            raise ValueError(f"필수 Sheet가 없습니다: {NODE_SHEET}")
        return _read_nodes(workbook[NODE_SHEET])
    finally:
        workbook.close()


def read_input_preview_from_excel(path_or_file) -> tuple[list[NodeRow], list[GuidewordRow]]:
    """업로드 Excel에서 화면 미리보기용 `#1`, `#2` 데이터를 읽습니다.

    쉽게 말하면 사용자가 AI 초안생성을 누르기 전에, Excel에서 읽힌 Node와
    각 Node에 연결된 변수/Guideword 조합을 화면에 먼저 보여주기 위한 함수입니다.
    """

    workbook = load_workbook(path_or_file, read_only=True, data_only=True)
    try:
        missing = [name for name in [NODE_SHEET, GUIDEWORD_SHEET] if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"필수 Sheet가 없습니다: {', '.join(missing)}")
        return _read_nodes(workbook[NODE_SHEET]), _read_guidewords(workbook[GUIDEWORD_SHEET])
    finally:
        workbook.close()


def validate_and_parse_excel(path: Path) -> tuple[list[NodeRow], list[GuidewordRow]]:
    """기존 호출부 호환을 위해 `#1`, `#2` 결과만 반환합니다."""

    nodes, guidewords, _criteria = validate_and_parse_excel_with_criteria(path)
    return nodes, guidewords


def validate_and_parse_excel_with_criteria(
    path: Path,
) -> tuple[list[NodeRow], list[GuidewordRow], RiskCriteria]:
    """업로드 Excel의 입력 Row와 위험도 기준표를 검증·구조화합니다.

    `#3`, `#4`는 AI가 채울 답안지라서 비어 있어도 정상입니다.
    이 함수는 사용자가 미리 작성해야 하는 "문제지"인 `#1`, `#2`만 검사합니다.
    """

    workbook = load_workbook(path)
    try:
        missing = [name for name in [NODE_SHEET, GUIDEWORD_SHEET] if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"필수 Sheet가 없습니다: {', '.join(missing)}")

        nodes = _read_nodes(workbook[NODE_SHEET])
        guidewords = _read_guidewords(workbook[GUIDEWORD_SHEET])

        node_keys = {(node.node_order, node.node_name) for node in nodes}
        invalid = [
            f"{row.node_order}/{row.node_name}"
            for row in guidewords
            if (row.node_order, row.node_name) not in node_keys
        ]
        if invalid:
            raise ValueError(f"#2 가이드워드에 #1 노드리스트와 맞지 않는 Node가 있습니다: {', '.join(invalid)}")

        if CRITERIA_SHEET in workbook.sheetnames:
            criteria = _read_criteria(workbook[CRITERIA_SHEET])
        else:
            criteria = _default_risk_criteria()

        return nodes, guidewords, criteria
    finally:
        workbook.close()


def export_result_excel(
    source_excel: Path,
    output_excel: Path,
    risk_rows: list[RiskAssessmentRow],
    action_rows: list[ActionPlanRow],
) -> None:
    """업로드 원본 Excel을 복사하듯 열고 `#3`, `#4` Sheet에 결과를 씁니다."""

    workbook = load_workbook(source_excel)
    _ensure_sheet_with_headers(workbook, RISK_SHEET, RISK_HEADERS)
    _ensure_sheet_with_headers(workbook, ACTION_SHEET, ACTION_HEADERS)

    risk_sheet = workbook[RISK_SHEET]
    action_sheet = workbook[ACTION_SHEET]
    _clear_data_rows(risk_sheet)
    _clear_data_rows(action_sheet)

    for row in risk_rows:
        risk_sheet.append(
            [
                row.no,
                row.node_order,
                row.node_name,
                row.parameter,
                row.guideword,
                row.deviation,
                row.cause,
                row.consequence,
                row.existing_safeguard,
                row.frequency,
                row.severity,
                row.risk_score,
                row.risk_level,
                row.action_required,
                _join_evidence(row.decision_evidence),
                _join_evidence(row.severity_evidence),
                _join_evidence(row.frequency_evidence),
                row.note,
            ]
        )

    for row in action_rows:
        action_sheet.append(
            [
                row.no,
                row.risk_assessment_no,
                row.node_name,
                row.recommendation,
                row.after_frequency,
                row.after_severity,
                row.after_risk_score,
                _join_evidence(row.evidence),
                row.note,
            ]
        )

    _style_sheet(risk_sheet)
    _style_sheet(action_sheet)
    workbook.save(output_excel)


def create_sample_excel(path: Path, nodes: list[dict[str, Any]], guidewords: list[dict[str, Any]]) -> None:
    """PoC용 샘플 Excel을 생성합니다.

    사용자의 요구대로 `#3`, `#4`는 헤더만 만들고 데이터는 비워둡니다.
    """

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    node_sheet = workbook.create_sheet(NODE_SHEET)
    guideword_sheet = workbook.create_sheet(GUIDEWORD_SHEET)
    risk_sheet = workbook.create_sheet(RISK_SHEET)
    action_sheet = workbook.create_sheet(ACTION_SHEET)
    criteria_sheet = workbook.create_sheet(CRITERIA_SHEET)

    _write_table(node_sheet, NODE_HEADERS, [[item["node_order"], item["node_name"]] for item in nodes])
    _write_table(
        guideword_sheet,
        GUIDEWORD_HEADERS,
        [[item["node_order"], item["node_name"], item["parameter"], item["guideword"]] for item in guidewords],
    )
    _write_table(risk_sheet, RISK_HEADERS, [])
    _write_table(action_sheet, ACTION_HEADERS, [])
    _write_criteria(criteria_sheet)

    for sheet in workbook.worksheets:
        _style_sheet(sheet)

    workbook.save(path)


def _read_nodes(sheet) -> list[NodeRow]:
    header_map = _header_map(sheet)
    _require_headers(header_map, NODE_HEADERS, NODE_SHEET)

    rows: list[NodeRow] = []
    for row_index in range(2, sheet.max_row + 1):
        order = sheet.cell(row_index, header_map["노드순서"]).value
        name = sheet.cell(row_index, header_map["노드명"]).value
        if order is None and name is None:
            continue
        rows.append(NodeRow(node_order=int(order), node_name=str(name).strip()))
    if not rows:
        raise ValueError("#1 노드리스트에 데이터가 없습니다.")
    return rows


def _read_guidewords(sheet) -> list[GuidewordRow]:
    header_map = _header_map(sheet)
    _require_headers(header_map, GUIDEWORD_HEADERS, GUIDEWORD_SHEET)

    rows: list[GuidewordRow] = []
    for row_index in range(2, sheet.max_row + 1):
        order = sheet.cell(row_index, header_map["노드순서"]).value
        node_name = sheet.cell(row_index, header_map["노드명"]).value
        parameter = sheet.cell(row_index, header_map["변수"]).value
        guideword = sheet.cell(row_index, header_map["가이드워드"]).value
        if all(value is None for value in [order, node_name, parameter, guideword]):
            continue
        rows.append(
            GuidewordRow(
                node_order=int(order),
                node_name=str(node_name).strip(),
                parameter=str(parameter).strip(),
                guideword=str(guideword).strip(),
            )
        )
    if not rows:
        raise ValueError("#2 가이드워드에 데이터가 없습니다.")
    return rows


def _read_criteria(sheet) -> RiskCriteria:
    """`위험도기준` Sheet를 Agent가 그대로 참고할 수 있는 구조로 읽습니다."""

    header_map = _header_map(sheet)
    _require_headers(header_map, CRITERIA_HEADERS, CRITERIA_SHEET)
    items: list[RiskCriterion] = []
    for row_index in range(2, sheet.max_row + 1):
        category = sheet.cell(row_index, header_map["구분"]).value
        score = sheet.cell(row_index, header_map["점수"]).value
        description = sheet.cell(row_index, header_map["기준"]).value
        if all(value is None for value in [category, score, description]):
            continue
        items.append(
            RiskCriterion(
                category=str(category).strip(),
                score=str(score).strip(),
                description=str(description).strip(),
            )
        )
    if not items:
        raise ValueError("위험도기준 Sheet에 데이터가 없습니다.")
    _validate_criteria_coverage(items)
    return RiskCriteria(items=items, source="업로드 Excel/위험도기준", requires_confirmation=False)


def _default_risk_criteria() -> RiskCriteria:
    return RiskCriteria(
        items=[
            RiskCriterion(category=str(category), score=str(score), description=str(description))
            for category, score, description in DEFAULT_CRITERIA_ROWS
        ],
        source="프로젝트 기본 위험도기준(업로드 Sheet 없음)",
        requires_confirmation=True,
    )


def _validate_criteria_coverage(items: list[RiskCriterion]) -> None:
    frequency_scores = {item.score for item in items if item.category == "빈도"}
    severity_scores = {item.score for item in items if item.category == "강도"}
    if frequency_scores != {"1", "2", "3", "4", "5"}:
        raise ValueError("위험도기준 Sheet의 빈도 점수는 1~5가 각각 한 번씩 필요합니다.")
    if severity_scores != {"1", "2", "3", "4"}:
        raise ValueError("위험도기준 Sheet의 강도 점수는 1~4가 각각 한 번씩 필요합니다.")


def _header_map(sheet) -> dict[str, int]:
    return {str(cell.value).strip(): cell.column for cell in sheet[1] if cell.value}


def _require_headers(header_map: dict[str, int], required: list[str], sheet_name: str) -> None:
    missing = [header for header in required if header not in header_map]
    if missing:
        raise ValueError(f"{sheet_name} Sheet에 필수 컬럼이 없습니다: {', '.join(missing)}")


def _write_table(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _ensure_sheet_with_headers(workbook, name: str, headers: list[str]) -> None:
    if name not in workbook.sheetnames:
        sheet = workbook.create_sheet(name)
        sheet.append(headers)
    else:
        sheet = workbook[name]
        if sheet.max_row == 0:
            sheet.append(headers)
        else:
            current = [cell.value for cell in sheet[1]]
            if current[: len(headers)] != headers:
                _clear_sheet(sheet)
                sheet.append(headers)


def _clear_sheet(sheet) -> None:
    if sheet.max_row:
        sheet.delete_rows(1, sheet.max_row)


def _clear_data_rows(sheet) -> None:
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)


def _write_criteria(sheet) -> None:
    sheet.append(CRITERIA_HEADERS)
    for row in DEFAULT_CRITERIA_ROWS:
        sheet.append(row)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 50)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _join_evidence(evidence) -> str:
    return "\n".join(f"- {item.reason} ({item.source})" for item in evidence)
