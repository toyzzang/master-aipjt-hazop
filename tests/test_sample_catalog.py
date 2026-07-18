import asyncio

from openpyxl import load_workbook

from app.hazop_engine.context import HazopDraftContext
from app.hazop_engine.workflow import generate_hazop_draft
from app.schemas.hazop import HazopInput
from app.services.excel import ACTION_SHEET, RISK_SHEET, validate_and_parse_excel
from app.services.msds import LOCAL_MSDS, material_names
from scripts.create_sample_excels import SAMPLE_DEFINITIONS, SAMPLE_DIR


def test_all_sample_excels_match_catalog_and_keep_output_sheets_empty():
    """모든 샘플이 입력만 포함하고 #3/#4 답안지는 비어 있는지 확인합니다."""

    assert len(SAMPLE_DEFINITIONS) >= 8

    for sample in SAMPLE_DEFINITIONS:
        path = SAMPLE_DIR / sample["filename"]
        assert path.exists(), sample["filename"]

        nodes, guidewords = validate_and_parse_excel(path)
        assert len(nodes) == len(sample["nodes"]), sample["filename"]
        assert len(guidewords) == len(sample["guidewords"]), sample["filename"]

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            assert workbook[RISK_SHEET].max_row == 1, sample["filename"]
            assert workbook[ACTION_SHEET].max_row == 1, sample["filename"]
        finally:
            workbook.close()


def test_complex_sample_is_materially_larger_than_legacy_samples():
    complex_sample = next(
        sample for sample in SAMPLE_DEFINITIONS if sample["filename"] == "HAZOP_Integrated_MultiUtility-Complex.xlsx"
    )

    assert len(complex_sample["nodes"]) >= 8
    assert len(complex_sample["guidewords"]) >= 20

    combinations = {
        (item["node_order"], item["node_name"], item["parameter"], item["guideword"])
        for item in complex_sample["guidewords"]
    }
    assert len(combinations) == len(complex_sample["guidewords"])


def test_demo_engine_processes_every_sample_without_inventing_rows(monkeypatch):
    """연결된 LLM이 없어도 8개 샘플 전체가 규칙 기반 초안을 끝까지 만드는지 확인합니다."""

    for name in [
        "AZURE_OPENAI_ENDPOINT",
        "AZURE_OPENAI_API_KEY",
        "AZURE_OPENAI_API_VERSION",
        "AZURE_OPENAI_DEPLOYMENT",
    ]:
        monkeypatch.delenv(name, raising=False)

    for sample in SAMPLE_DEFINITIONS:
        path = SAMPLE_DIR / sample["filename"]
        nodes, guidewords = validate_and_parse_excel(path)
        names = material_names(sample["materials"])
        msds_context = {
            name.lower(): LOCAL_MSDS.get(name.lower())
            for name in names
            if LOCAL_MSDS.get(name.lower()) is not None
        }
        context = HazopDraftContext(
            input_data=HazopInput(
                maker="Sample",
                model=sample["filename"],
                materials=sample["materials"],
                notes=sample["purpose"],
            ),
            nodes=nodes,
            guidewords=guidewords,
            msds_context=msds_context,
        )

        result = asyncio.run(generate_hazop_draft(context))

        assert result.mode == "demo", sample["filename"]
        assert len(result.risk_rows) == len(guidewords), sample["filename"]
        allowed = {(row.node_order, row.node_name, row.parameter, row.guideword) for row in guidewords}
        generated = {(row.node_order, row.node_name, row.parameter, row.guideword) for row in result.risk_rows}
        assert generated == allowed, sample["filename"]

        high_risk_numbers = {row.no for row in result.risk_rows if row.risk_score >= 9}
        assert {row.risk_assessment_no for row in result.action_rows} == high_risk_numbers, sample["filename"]
