import asyncio
import json
from pathlib import Path

from openpyxl import load_workbook

from app.hazop_engine.context import HazopDraftContext
from app.hazop_engine.workflow import generate_hazop_draft
from app.schemas.hazop import HazopInput
from app.services.excel import ACTION_SHEET, RISK_SHEET, export_result_excel, validate_and_parse_excel_with_criteria
from app.services.msds import MsdsSummary


BASE_DIR = Path(__file__).resolve().parent.parent
SCENARIO_PATH = BASE_DIR / "tests" / "scenarios" / "hazop_engine_scenarios.json"


def test_hazop_engine_reusable_scenarios(monkeypatch, tmp_path):
    """저장된 HAZOP 시나리오 기준으로 Engine 결과와 Excel export를 검증합니다."""

    _disable_connected_model(monkeypatch)
    scenarios = json.loads(SCENARIO_PATH.read_text(encoding="utf-8"))

    for scenario in scenarios:
        sample_excel = BASE_DIR / scenario["sample_excel"]
        nodes, guidewords, risk_criteria = validate_and_parse_excel_with_criteria(sample_excel)
        expected = scenario["expected"]

        assert len(nodes) == expected["node_count"], scenario["id"]
        assert len(guidewords) == expected["guideword_count"], scenario["id"]
        assert len(risk_criteria.items) == 15, scenario["id"]
        assert risk_criteria.source == "업로드 Excel/위험도기준", scenario["id"]
        assert not risk_criteria.requires_confirmation, scenario["id"]

        context = HazopDraftContext(
            input_data=HazopInput(**scenario["input"]),
            nodes=nodes,
            guidewords=guidewords,
            risk_criteria=risk_criteria,
            msds_context=_msds_context(scenario),
        )
        result = asyncio.run(generate_hazop_draft(context))

        assert result.mode == "demo", scenario["id"]
        assert len(result.risk_rows) == expected["risk_count"], scenario["id"]
        assert len(result.action_rows) == expected["action_count"], scenario["id"]

        allowed = {(item.node_order, item.node_name, item.parameter, item.guideword) for item in guidewords}
        generated = {(row.node_order, row.node_name, row.parameter, row.guideword) for row in result.risk_rows}
        assert generated == allowed, scenario["id"]

        high_risk_rows = [row for row in result.risk_rows if row.risk_score >= 9]
        assert len(high_risk_rows) == expected["high_risk_count"], scenario["id"]
        assert {row.risk_assessment_no for row in result.action_rows} == {row.no for row in high_risk_rows}, scenario["id"]

        for row in result.risk_rows:
            assert 1 <= row.frequency <= 5, scenario["id"]
            assert 1 <= row.severity <= 4, scenario["id"]
            assert row.risk_score == row.frequency * row.severity, scenario["id"]
            assert row.decision_evidence, scenario["id"]
            assert row.severity_evidence, scenario["id"]
            assert row.frequency_evidence, scenario["id"]
            assert any("사고이력" in evidence.reason or "표준 HAZOP" in evidence.reason for evidence in row.frequency_evidence), scenario["id"]

        output_excel = tmp_path / f"{scenario['id']}_result.xlsx"
        export_result_excel(sample_excel, output_excel, result.risk_rows, result.action_rows)
        assert output_excel.exists(), scenario["id"]
        _assert_output_excel_row_counts(output_excel, expected["risk_count"], expected["action_count"], scenario["id"])


def test_missing_criteria_sheet_uses_default_and_requires_confirmation(tmp_path):
    source = BASE_DIR / "samples" / "HAZOP_CleanTech_CT-DIW-100.xlsx"
    workbook = load_workbook(source)
    del workbook["위험도기준"]
    target = tmp_path / "without-criteria.xlsx"
    workbook.save(target)
    workbook.close()

    _nodes, _guidewords, criteria = validate_and_parse_excel_with_criteria(target)

    assert len(criteria.items) == 15
    assert criteria.requires_confirmation
    assert "업로드 Sheet 없음" in criteria.source


def _disable_connected_model(monkeypatch):
    for name in [
        "AZURE_OPENAI_ENDPOINT",
        "AZURE_OPENAI_API_KEY",
        "AZURE_OPENAI_API_VERSION",
        "AZURE_OPENAI_DEPLOYMENT",
    ]:
        monkeypatch.delenv(name, raising=False)


def _msds_context(scenario: dict) -> dict[str, MsdsSummary]:
    return {key: MsdsSummary(**value) for key, value in scenario["msds"].items()}


def _assert_output_excel_row_counts(path: Path, risk_count: int, action_count: int, scenario_id: str) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert workbook[RISK_SHEET].max_row == risk_count + 1, scenario_id
        assert workbook[ACTION_SHEET].max_row == action_count + 1, scenario_id
    finally:
        workbook.close()
